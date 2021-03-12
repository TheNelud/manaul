import openpyxl
#Подключаем книгу
book = openpyxl.open("Оценка ОПФ.xlsx", read_only=True)
sheet_first = book.active #Лист первый

# sheet_second = book.worksheets[2] #Выбираем второй лист
# print(sheet_second['A2'])

# sheets = book.worksheets #Список листов
# print(sheets)


#вытаскиваем отдельную ячейку
# a1 = (sheet_first["A1"].value)
# print(a1)
#вытаскиваем отдельную ячейку с помощью индекса [строка][столбец] B1
# b1 = (sheet_first[1][1].value)          #Колонки начинаются с 0, а ряды с 1
# print(b1)

#Обращение к интересующие нас ряды
# for row in range(1, 10):
#     print(sheet_first[row])

#Обращение к интересующие нас рядах
#for row in range(1, 10):
#    first_column = sheet_first[row][0].value
#    second_column = sheet_first[row][1].value
#    theard_column = sheet_first[row][2].value
#    fourth_column = sheet_first[row][3].value
#    fiveth_column = sheet_first[row][4].value
#    print(first_column,second_column,theard_column,fourth_column,fiveth_column)

#Обращение ко всем рядам
# for row in range(1,sheet_first.max_row+1):
#     first_column = sheet_first[row][0].value
#     second_column = sheet_first[row][1].value
#     theard_column = sheet_first[row][2].value
#     fourth_column = sheet_first[row][3].value
#     fiveth_column = sheet_first[row][4].value
#     print(row, first_column,second_column,theard_column,fourth_column,fiveth_column)

#Работаем с диапазоном
#cells = sheet_first['B1' : 'C11']
# for cell in cells:
#     print(cell)

# for second_column, theard_column in cells:
#     print(second_column.value, theard_column.value)

# for row in sheet_first.iter_rows(min_row=2, max_row=20, min_col=1, max_col=3):
#     #print(row)
#     for cell in row:
#         print(cell.value, end= ' ')
#     print(" ")

#Обходим весь файл
# for row in sheet_first.iter_rows():
#     for cell in row:
#         print(cell.value, end= ' ')
#     print(" ")